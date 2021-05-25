import openpyxl, pprint

wb = openpyxl.load_workbook("censuspopdata.xlsx")

sheet = wb.active
population_dict = {}
county_list = []
# ~ pp = pprint.PrettyPrinter()

for i in range(2,sheet.max_row+1):
    if sheet.cell(row = i, column = 3).value not in county_list:
        county_list.append(sheet.cell(row = i, column = 3).value)

for i in range(0,len(county_list)):
    population_dict[county_list[i]] = 0

for i in range(2,sheet.max_row+1):
    population_dict[sheet.cell(row = i, column = 3).value] = population_dict[sheet.cell(row = i, column = 3).value] + sheet.cell(row = i, column = 4).value


with open("results.txt","w") as f:
    f.write(pprint.pformat(population_dict))
