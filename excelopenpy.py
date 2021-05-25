import openpyxl

wb = openpyxl.load_workbook("example.xlsx")

# ~ sheetlist = wb.sheetnames # returns a list of sheets
# ~ sheet = wb["Sheet3"]
# ~ print(sheet.title)
# ~ active = wb.active
# ~ print(active.title)

sheet = wb["Sheet1"]
# ~ print(f"A1: {sheet['A1'].value}")
# ~ print(f"B1: {sheet['B1'].value}")
# ~ print(f"C1: {sheet['C1'].value}")

# ~ print(f"type A1: {type(sheet['A1'].value)}")
# ~ print(f"type B1: {type(sheet['B1'].value)}")
# ~ print(f"type C1: {type(sheet['C1'].value)}")

# ~ for i in range(1,8):
   # ~ if sheet.cell(row=i, column=2).value == "Bananas":
       # ~ print(f"Row: {i}")

# ~ print(sheet.max_row)
# ~ print(sheet.max_column)

# ~ for i in sheet["A1":"C3"]:
    # ~ for j in i:
        # ~ print(j.coordinate, j.value)

for i in range(1, sheet.max_row+1):
    print(f"i: {i} {sheet.cell(row = i, column = 2).value}")

